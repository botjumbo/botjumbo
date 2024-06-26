import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoSuchWindowException
from datetime import datetime
import openpyxl
import xlwings as xw
import re
import time


fecha_actual = datetime.now().strftime('%d-%m-%Y')

# Guarda el tiempo de inicio
inicio = time.time()

urls = [
    #Pan y Cereales
    
    #Carnes y derivados
    
    #Pescados y Mariscos
    
    #Leche , productos lacteos y huevos
    
    #Aceites, grasas y manteca
    #Frutas
    
    #Verduras, tuberculos y legumbres
    
    #Azúcar, dulces, chocolates y golosinas
    
    #Otros Alimentos
    
    #Bebidas no Alcoholicas    
        #Cafe , Te, Yerba, Cacao
    
        #Aguas minerales, bebidas gaseosas y jugos
    
]

titulos_precios = []
workbook = openpyxl.load_workbook("DatosJumbo1.xlsx")
sheet = workbook["Diaria"]
fila_vacia = 1
columna = 2 
fila_titulos = 1
columna_titulos = 2

# Especifica la ruta del perfil de Chrome personalizado
profile_directory = 'C:\\Users\\#Nombre de pc#\\AppData\\Local\\Google\\Chrome\\User Data\\'

# Configurar el driver de Selenium con el perfil personalizado
options = webdriver.ChromeOptions()
options.add_argument(f"user-data-dir={profile_directory}")
service = Service()
driver = webdriver.Chrome(service=service, options=options)

# Iterar sobre las URLs
for url in urls:
    try:
    #Averiguar si hay stock
        driver.get(url)
        time.sleep(4)
        titulo = driver.find_element(By.XPATH, "//h1[@class='vtex-store-components-3-x-productNameContainer mv0 t-heading-4']/span[@class='vtex-store-components-3-x-productBrand ']").text.strip()
        stock  = driver.find_element(By.XPATH, "//div[contains(@class, 'vtex-flex-layout-0-x-flexColChild')]//p[contains(@class, 'vtex-outOfStockFlag__text')]").text.strip() 
        if stock == "Producto sin stock":
            titulos_precios.append((titulo,0)) #si no hay stock agrego el precio como 0
            print(titulo,0)
    except NoSuchElementException:
        try:
            # Obtener el titulo
            titulo = driver.find_element(By.XPATH, "//h1[@class='vtex-store-components-3-x-productNameContainer mv0 t-heading-4']/span[@class='vtex-store-components-3-x-productBrand ']").text.strip()
        except NoSuchElementException:
            print("No se encontró el Titulo en la página")
            titulos_precios.append(("Producto X", 0))
            continue
        try:
            #Obtener el precio
            #precio de lista
            precios = driver.find_elements(By.XPATH, "//div[contains(@class, 'vtex-flex-layout-0-x-flexColChild') and contains(@class, 'vtex-flex-layout-0-x-flexColChild--separator') and contains(@class, 'vtex-flex-layout-0-x-flexColChild--product-box') and contains(@class, 'pb0')]//span[contains(@class, 'jumboargentinaio-store-theme-1QiyQadHj-1_x9js9EXUYK')]")

            #precio con descuentos
            #precios = driver.find_elements(By.XPATH, "//div[contains(@class, 'vtex-flex-layout-0-x-flexColChild--separator') and .//div[contains(text(), 'espacio')]]//div[contains(@class, 'jumboargentinaio-store-theme-1dCOMij_MzTzZOCohX1K7w')]")

            for precio in precios:
                precio_texto = precio.get_attribute("innerText")

        except (NoSuchElementException,NoSuchWindowException) as e:
                
                titulos_precios.append((titulo,0))
                print(titulo, 0)
                continue

        if precio_texto:
            #Si encuentra precios por el primer XPATH de precios
            try:

                precio = re.search(r'(?<=\$)[\d,.]+', precio_texto)
                precio = precio.group().replace('.','').replace(',','.')
                precio = float(precio)

        
                print(titulo,precio)
                titulos_precios.append((titulo,precio))
                continue
            except Exception as e:
                titulos_precios.append((titulo,0))
                print(titulo, 0)
                continue



driver.quit()


# Guarda el tiempo de finalización
fin = time.time()


# Calcula la duración total
duracion = round((fin - inicio),2)
min = round((duracion/60),2)

# Imprime la duración en segundos
print(f"El programa tardó {min} minutos en ejecutarse.")


while sheet[f"A{fila_vacia}"].value is not None:
    fila_vacia += 1

#agregar fecha
sheet[f"A{fila_vacia}"] = fecha_actual


for _, precio in titulos_precios:
    sheet.cell(row=fila_vacia, column=columna).value = precio
    columna += 1  

for  titulo, _ in titulos_precios:

    sheet.cell(row=fila_titulos, column=columna_titulos).value = titulo
    columna_titulos += 1  



# Guardar el libro de trabajo datos
workbook.save("DatosJumbo1.xlsx") 
